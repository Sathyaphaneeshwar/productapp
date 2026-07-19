"""Validate, preview, and commit user-managed stock records."""

from __future__ import annotations

import csv
import io
import json
import os
import re
import time
from pathlib import Path
from typing import Any

from config import DATABASE_PATH
from db import get_db_connection


MAX_FILE_BYTES = 5 * 1024 * 1024
MAX_ROWS = 25_000
ALLOWED_SUFFIXES = {".csv", ".tsv", ".xlsx"}
ISIN_PATTERN = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$")


class StockImportError(ValueError):
    def __init__(
        self,
        message: str,
        status_code: int = 400,
        *,
        details: dict[str, Any] | None = None,
    ):
        super().__init__(message)
        self.status_code = status_code
        self.details = details or {}


def _header_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


HEADER_ALIASES = {
    "isin": "isin",
    "isinnumber": "isin",
    "isinno": "isin",
    "companyname": "company_name",
    "name": "company_name",
    "stockname": "company_name",
    "nameofcompany": "company_name",
    "issuername": "company_name",
}


def normalize_isin(value: Any) -> str:
    return str(value or "").strip().upper()


def is_valid_isin(value: Any) -> bool:
    """Validate the 12-character format and ISO 6166 Luhn check digit."""
    isin = normalize_isin(value)
    if not ISIN_PATTERN.fullmatch(isin):
        return False

    expanded = "".join(
        str(ord(character) - 55) if character.isalpha() else character
        for character in isin
    )
    total = 0
    should_double = False
    for character in reversed(expanded):
        digit = int(character)
        if should_double:
            digit *= 2
            if digit > 9:
                digit -= 9
        total += digit
        should_double = not should_double
    return total % 10 == 0


class StockImportService:
    def __init__(self, db_path: str | Path = DATABASE_PATH):
        self.db_path = str(db_path)

    def get_db_connection(self):
        return get_db_connection(self.db_path)

    @staticmethod
    def _map_headers(headers: list[Any]) -> dict[int, str]:
        mapped: dict[int, str] = {}
        for index, header in enumerate(headers):
            canonical = HEADER_ALIASES.get(_header_key(header))
            if canonical and canonical not in mapped.values():
                mapped[index] = canonical
        missing = {"isin", "company_name"} - set(mapped.values())
        if missing:
            raise StockImportError(
                "Header row must contain exactly the required fields "
                "ISIN and CompanyName."
            )
        return mapped

    @classmethod
    def _rows_from_matrix(cls, matrix: list[list[Any]]) -> list[dict[str, Any]]:
        if not matrix:
            raise StockImportError("The uploaded file is empty.")
        mapping = cls._map_headers(matrix[0])
        rows: list[dict[str, Any]] = []
        for row_number, values in enumerate(matrix[1:], start=2):
            record = {
                canonical: values[index] if index < len(values) else None
                for index, canonical in mapping.items()
            }
            if not any(str(value or "").strip() for value in record.values()):
                continue
            record["row"] = row_number
            rows.append(record)
            if len(rows) > MAX_ROWS:
                raise StockImportError(
                    f"Files may contain at most {MAX_ROWS:,} stock rows.",
                    status_code=413,
                )
        return rows

    @classmethod
    def parse_file(cls, filename: str, content: bytes) -> list[dict[str, Any]]:
        suffix = Path(filename or "").suffix.lower()
        if suffix not in ALLOWED_SUFFIXES:
            if suffix == ".xls":
                raise StockImportError(
                    "Old .xls files are not supported. Save the file as .xlsx or .csv.",
                    status_code=415,
                )
            raise StockImportError(
                "Upload a .csv, .tsv, or .xlsx file.",
                status_code=415,
            )
        if len(content) > MAX_FILE_BYTES:
            raise StockImportError("The upload exceeds the 5 MB limit.", status_code=413)
        if not content:
            raise StockImportError("The uploaded file is empty.")

        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as error:
                raise StockImportError(
                    "XLSX support is unavailable in this installation."
                ) from error
            try:
                workbook = load_workbook(
                    io.BytesIO(content),
                    read_only=True,
                    data_only=True,
                )
                worksheet = workbook.worksheets[0]
                matrix = [list(row) for row in worksheet.iter_rows(values_only=True)]
            except Exception as error:
                raise StockImportError(f"Could not read the XLSX file: {error}") from error
            finally:
                if "workbook" in locals():
                    workbook.close()
            return cls._rows_from_matrix(matrix)

        decoded = None
        for encoding in ("utf-8-sig", "cp1252"):
            try:
                decoded = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            raise StockImportError("The text file encoding is not supported.")

        delimiter = "\t" if suffix == ".tsv" else ","
        reader = csv.reader(io.StringIO(decoded), delimiter=delimiter)
        return cls._rows_from_matrix([list(row) for row in reader])

    @staticmethod
    def validate_rows(
        rows: list[dict[str, Any]],
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
        valid_by_isin: dict[str, dict[str, Any]] = {}
        invalid: list[dict[str, Any]] = []
        warnings: list[dict[str, Any]] = []

        for raw in rows:
            row_number = int(raw.get("row") or 0)
            isin = normalize_isin(raw.get("isin"))
            company_name = " ".join(str(raw.get("company_name") or "").split())
            data = {"ISIN": isin, "CompanyName": company_name}

            if not is_valid_isin(isin):
                invalid.append(
                    {
                        "row": row_number,
                        "reason": "ISIN is invalid or has an incorrect check digit.",
                        "data": data,
                    }
                )
                continue
            if not company_name:
                invalid.append(
                    {
                        "row": row_number,
                        "reason": "CompanyName is required.",
                        "data": data,
                    }
                )
                continue

            if isin in valid_by_isin:
                warnings.append(
                    {
                        "row": row_number,
                        "reason": f"Duplicate ISIN {isin}; the last row was used.",
                    }
                )
                del valid_by_isin[isin]
            valid_by_isin[isin] = {
                "row": row_number,
                "isin": isin,
                "company_name": company_name,
            }

        return list(valid_by_isin.values()), invalid, warnings

    @staticmethod
    def _existing_by_isin(connection, isins: list[str]) -> dict[str, dict[str, Any]]:
        existing: dict[str, dict[str, Any]] = {}
        for offset in range(0, len(isins), 500):
            batch = isins[offset:offset + 500]
            if not batch:
                continue
            placeholders = ",".join("?" for _ in batch)
            rows = connection.execute(
                f"""
                SELECT id, isin_number, stock_name, source, is_active
                FROM stocks
                WHERE isin_number IN ({placeholders})
                """,
                batch,
            ).fetchall()
            existing.update({row["isin_number"]: dict(row) for row in rows})
        return existing

    @classmethod
    def diff_rows(cls, connection, valid: list[dict[str, Any]]) -> dict[str, Any]:
        existing = cls._existing_by_isin(
            connection,
            [row["isin"] for row in valid],
        )
        new_rows = []
        updated_rows = []
        unchanged = 0
        for row in valid:
            current = existing.get(row["isin"])
            if current is None:
                new_rows.append(row)
                continue

            changes: dict[str, list[Any]] = {}
            if current["stock_name"] != row["company_name"]:
                changes["CompanyName"] = [
                    current["stock_name"],
                    row["company_name"],
                ]
            if not bool(current["is_active"]):
                changes["Active"] = [False, True]
            if changes:
                updated_rows.append({**row, "field_changes": changes})
            else:
                unchanged += 1
        return {
            "new_rows": new_rows,
            "updated_rows": updated_rows,
            "unchanged": unchanged,
        }

    def preview(
        self,
        filename: str,
        content: bytes,
    ) -> dict[str, Any]:
        parsed = self.parse_file(filename, content)
        valid, invalid, warnings = self.validate_rows(parsed)
        connection = self.get_db_connection()
        try:
            diff = self.diff_rows(connection, valid)
            safe_filename = os.path.basename(filename or "stocks.csv")[:255]
            cursor = connection.execute(
                """
                INSERT INTO stock_import_batches (
                    filename, uploaded_ts, rows_total, rows_new, rows_updated,
                    rows_unchanged, rows_invalid, status, payload_json,
                    warnings_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, 'previewed', ?, ?)
                """,
                (
                    safe_filename,
                    int(time.time()),
                    len(parsed),
                    len(diff["new_rows"]),
                    len(diff["updated_rows"]),
                    diff["unchanged"],
                    len(invalid),
                    json.dumps(valid),
                    json.dumps(warnings),
                ),
            )
            connection.commit()
            batch_id = cursor.lastrowid
        finally:
            connection.close()

        return {
            "batch_id": batch_id,
            "filename": safe_filename,
            "rows_total": len(parsed),
            "new": len(diff["new_rows"]),
            "updated": len(diff["updated_rows"]),
            "unchanged": diff["unchanged"],
            "invalid_count": len(invalid),
            "invalid": invalid[:100],
            "warnings": warnings[:100],
            "sample_new": diff["new_rows"][:20],
            "sample_updated": diff["updated_rows"][:20],
        }

    def commit_batch(self, batch_id: int) -> dict[str, Any]:
        connection = self.get_db_connection()
        try:
            connection.execute("BEGIN IMMEDIATE")
            batch = connection.execute(
                """
                SELECT *
                FROM stock_import_batches
                WHERE id = ?
                """,
                (int(batch_id),),
            ).fetchone()
            if not batch:
                raise StockImportError("Import preview not found.", status_code=404)
            if batch["status"] != "previewed":
                raise StockImportError(
                    "This import preview has already been committed.",
                    status_code=409,
                )

            valid = json.loads(batch["payload_json"])
            if not valid:
                raise StockImportError(
                    "This preview has no valid stock rows to import.",
                    status_code=409,
                )
            diff = self.diff_rows(connection, valid)
            changed_isins = {
                row["isin"] for row in diff["new_rows"] + diff["updated_rows"]
            }
            for row in valid:
                connection.execute(
                    """
                    INSERT INTO stocks (
                        stock_symbol, bse_code, isin_number, stock_name,
                        source, is_active, created_at, updated_at
                    ) VALUES (NULL, NULL, ?, ?, 'import', 1,
                              CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                    ON CONFLICT(isin_number) DO UPDATE SET
                        stock_name = excluded.stock_name,
                        is_active = 1,
                        updated_at = CURRENT_TIMESTAMP
                    """,
                    (row["isin"], row["company_name"]),
                )

            has_activity_table = connection.execute(
                """
                SELECT 1 FROM sqlite_master
                WHERE type = 'table' AND name = 'stock_activity_logs'
                """
            ).fetchone()
            if has_activity_table:
                for row in valid:
                    if row["isin"] not in changed_isins:
                        continue
                    stock = connection.execute(
                        "SELECT id FROM stocks WHERE isin_number = ?",
                        (row["isin"],),
                    ).fetchone()
                    connection.execute(
                        """
                        INSERT INTO stock_activity_logs (
                            stock_id, stage, level, message, details_json,
                            created_at
                        ) VALUES (?, 'stock_import', 'success', ?, ?,
                                  CURRENT_TIMESTAMP)
                        """,
                        (
                            stock["id"],
                            "Stock imported or updated",
                            json.dumps(
                                {
                                    "filename": batch["filename"],
                                    "isin": row["isin"],
                                }
                            ),
                        ),
                    )

            committed_ts = int(time.time())
            connection.execute(
                """
                UPDATE stock_import_batches
                SET status = 'committed',
                    rows_new = ?,
                    rows_updated = ?,
                    rows_unchanged = ?,
                    committed_ts = ?
                WHERE id = ?
                """,
                (
                    len(diff["new_rows"]),
                    len(diff["updated_rows"]),
                    diff["unchanged"],
                    committed_ts,
                    int(batch_id),
                ),
            )
            connection.commit()
            return {
                "new": len(diff["new_rows"]),
                "updated": len(diff["updated_rows"]),
                "unchanged": diff["unchanged"],
            }
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()

    def create_manual(self, isin: Any, company_name: Any) -> dict[str, Any]:
        valid, invalid, _warnings = self.validate_rows(
            [{"row": 1, "isin": isin, "company_name": company_name}]
        )
        if invalid:
            raise StockImportError(invalid[0]["reason"])

        row = valid[0]
        connection = self.get_db_connection()
        try:
            connection.execute("BEGIN IMMEDIATE")
            existing = connection.execute(
                """
                SELECT id, isin_number AS isin, stock_name AS company_name,
                       source, is_active
                FROM stocks
                WHERE isin_number = ?
                """,
                (row["isin"],),
            ).fetchone()
            if existing:
                raise StockImportError(
                    "This ISIN already exists.",
                    status_code=409,
                    details={"existing": dict(existing)},
                )
            cursor = connection.execute(
                """
                INSERT INTO stocks (
                    stock_symbol, bse_code, isin_number, stock_name,
                    source, is_active, created_at, updated_at
                ) VALUES (NULL, NULL, ?, ?, 'manual', 1,
                          CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """,
                (row["isin"], row["company_name"]),
            )
            stock_id = cursor.lastrowid
            if connection.execute(
                """
                SELECT 1 FROM sqlite_master
                WHERE type = 'table' AND name = 'stock_activity_logs'
                """
            ).fetchone():
                connection.execute(
                    """
                    INSERT INTO stock_activity_logs (
                        stock_id, stage, level, message, details_json, created_at
                    ) VALUES (?, 'stock_import', 'success',
                              'Stock added manually', ?, CURRENT_TIMESTAMP)
                    """,
                    (stock_id, json.dumps({"isin": row["isin"]})),
                )
            connection.commit()
            return {
                "id": stock_id,
                "isin": row["isin"],
                "company_name": row["company_name"],
                "source": "manual",
                "is_active": True,
            }
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()

    def prune_stale_previews(self) -> int:
        connection = self.get_db_connection()
        try:
            cursor = connection.execute(
                """
                DELETE FROM stock_import_batches
                WHERE status = 'previewed' AND uploaded_ts < ?
                """,
                (int(time.time()) - 86_400,),
            )
            connection.commit()
            return cursor.rowcount
        finally:
            connection.close()
